#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""وب‌اپ حسابداری هلیا بیوتی — Flask v2"""
import os, sys, json, secrets, re
from datetime import datetime, timedelta
from collections import defaultdict
from flask import Flask, render_template, request, redirect, url_for, send_file, flash, jsonify, session, g
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import auth
import invoice_pdf

# login_required is provided by auth module
login_required = auth.login_required

# ─── Portable paths (work both as script and as PyInstaller EXE) ───
import sys as _sys
def _base_dir():
    # When frozen by PyInstaller, keep data next to the executable so it
    # persists and is easy to back up / migrate to a server later.
    if getattr(_sys, "frozen", False):
        return os.path.dirname(_sys.executable)
    return os.path.dirname(os.path.abspath(__file__))
BASE_DIR = _base_dir()
TEMPLATE_DIR = os.path.join(BASE_DIR, "templates") if getattr(_sys, "frozen", False) else "templates"
# Only set a static folder if it actually exists (avoids 500 when frozen without a static/ dir)
if getattr(_sys, "frozen", False):
    STATIC_DIR = os.path.join(BASE_DIR, "static") if os.path.isdir(os.path.join(BASE_DIR, "static")) else None
else:
    STATIC_DIR = "static" if os.path.isdir("static") else None

app = Flask(__name__, template_folder=TEMPLATE_DIR, static_folder=STATIC_DIR)
app.secret_key = "helia-beauty-salon-fixed-secret-key-2024"  # fixed for persistent sessions
app.config["SESSION_COOKIE_NAME"] = "helia_session"
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"

# Login is ENABLED by default. Set True only for quick local testing.
LOGIN_DISABLED = False
# Secure flag set per-request in before_request (depends on scheme)

# ─── Error logging (so EXE builds can be debugged via app_error.log) ───
import traceback as _tb
@app.errorhandler(500)
def _log_500(e):
    tb = _tb.format_exc()
    try:
        with open(os.path.join(BASE_DIR, "app_error.log"), "a", encoding="utf-8") as f:
            f.write("\n=== 500 at %s ===\n" % PersianDate.today_str())
            f.write(tb)
    except Exception:
        pass
    # Also print to console so it is visible in the EXE window
    try:
        print("!!! 500 ERROR !!!")
        print(tb)
    except Exception:
        pass
    return "Internal Server Error - details saved to app_error.log", 500

# ─── CSRF Protection (session-based, no extra deps) ───
def generate_csrf_token():
    if "csrf_token" not in session:
        session["csrf_token"] = secrets.token_hex(32)
    return session["csrf_token"]

@app.context_processor
def inject_csrf():
    return dict(csrf_token=generate_csrf_token(), auth=auth)

def csrf_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if request.method == "POST":
            token = session.get("csrf_token", "")
            submitted = request.form.get("csrf_token", "") or request.headers.get("X-CSRF-Token", "")
            if not token or not secrets.compare_digest(token, submitted):
                if request.path.startswith("/api/"):
                    return jsonify({"error": "CSRF token نامعتبر است"}), 400
                if request.endpoint == "login":
                    flash("درخواست نامعتبر است، دوباره تلاش کنید", "error")
                    return render_template("login.html")
                flash("درخواست نامعتبر است (CSRF)", "error")
                return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated

# ─── Rate Limiting (simple, SQLite-backed) ───
import sqlite3 as _sql
_RATE_DB = os.path.join(BASE_DIR, "data", "ratelimit.db")
def _rate_get_conn():
    conn = _sql.connect(_RATE_DB)
    conn.execute("""CREATE TABLE IF NOT EXISTS login_attempts (
        ip TEXT PRIMARY KEY, fails INTEGER DEFAULT 0,
        first_fail TEXT, locked_until TEXT)""")
    return conn

def check_rate_limit(ip):
    """Return (allowed: bool, retry_after_sec: int)."""
    conn = _rate_get_conn()
    now = datetime.now()
    row = conn.execute("SELECT fails, locked_until FROM login_attempts WHERE ip=?", (ip,)).fetchone()
    if row:
        locked = row[1]
        if locked:
            locked_dt = datetime.strptime(locked, "%Y-%m-%d %H:%M:%S")
            if now < locked_dt:
                secs = int((locked_dt - now).total_seconds())
                conn.close()
                return False, secs
    conn.close()
    return True, 0

def register_fail(ip):
    conn = _rate_get_conn()
    now = datetime.now()
    row = conn.execute("SELECT fails FROM login_attempts WHERE ip=?", (ip,)).fetchone()
    fails = (row[0] if row else 0) + 1
    if fails >= 5:
        lock = (now + timedelta(minutes=30)).strftime("%Y-%m-%d %H:%M:%S")
        conn.execute("""INSERT INTO login_attempts (ip, fails, locked_until) VALUES (?,?,?)
            ON CONFLICT(ip) DO UPDATE SET fails=?, locked_until=?""",
            (ip, fails, lock, fails, lock))
    else:
        conn.execute("""INSERT INTO login_attempts (ip, fails, first_fail) VALUES (?,?,?)
            ON CONFLICT(ip) DO UPDATE SET fails=?""",
            (ip, fails, now.strftime("%Y-%m-%d %H:%M:%S"), fails))
    conn.commit(); conn.close()

def reset_fails(ip):
    conn = _rate_get_conn()
    conn.execute("DELETE FROM login_attempts WHERE ip=?", (ip,))
    conn.commit(); conn.close()

@app.before_request
def require_login():
    # TEMPORARY: when login disabled (for testing), allow all pages
    if LOGIN_DISABLED:
        # still set a default role so RBAC/templates don't break
        if "role" not in session:
            session["role"] = "admin"
            session["user"] = "test"
        return None
    # Set Secure cookie flag if served over HTTPS (tunnel)
    if request.scheme == "https":
        app.config["SESSION_COOKIE_SECURE"] = True
    # Public routes that don't need auth
    public = {"login", "logout", "static"}
    if request.endpoint in public or request.endpoint is None:
        return None
    if "user" not in session:
        # Don't redirect API calls to HTML login; return 401 JSON
        if request.path.startswith("/api/"):
            return jsonify({"error": "احراز هویت لازم است"}), 401
        return redirect(url_for("login"))
    # Role-based access control
    role = session.get("role")
    path = request.path
    if not auth.role_can_access(role, path):
        # Avoid redirect loop: send to first page the role CAN access
        if path in ("/dashboard",):  # already there but no access -> go to first allowed
            first_allowed = next((r for r in ["/", "/customers", "/reports", "/login"] if auth.role_can_access(role, r)), "/login")
            return redirect(first_allowed)
        flash("شما دسترسی لازم برای این بخش را ندارید", "error")
        first_allowed = next((r for r in ["/customers", "/", "/reports", "/dashboard"] if auth.role_can_access(role, r)), "/login")
        return redirect(first_allowed)
    return None

# ─── Data Directory (portable: next to exe when frozen) ───
DATA_DIR = os.path.join(BASE_DIR, "data")
os.makedirs(DATA_DIR, exist_ok=True)
EMPLOYEES_FILE = os.path.join(DATA_DIR, "employees.xlsx")
SERVICES_FILE = os.path.join(DATA_DIR, "services.xlsx")
TRANSACTIONS_FILE = os.path.join(DATA_DIR, "transactions.xlsx")
CUSTOMERS_FILE = os.path.join(DATA_DIR, "customers.xlsx")
EXPENSES_FILE = os.path.join(DATA_DIR, "expenses.xlsx")
INVENTORY_FILE = os.path.join(DATA_DIR, "inventory.xlsx")
INVENTORY_LOG_FILE = os.path.join(DATA_DIR, "inventory_log.xlsx")

# ─── Persian Date ───
class PersianDate:
    @staticmethod
    def gregorian_to_jalali(gy, gm, gd):
        g_d_m = [0, 31, 59, 90, 120, 151, 181, 212, 243, 273, 304, 334]
        gy2 = gy + 1 if gm > 2 else gy
        days = 355666 + (365*gy) + ((gy2+3)//4) - ((gy2+99)//100) + ((gy2+399)//400) + gd + g_d_m[gm-1]
        jy = -1595 + (33*(days//12053)); days %= 12053
        jy += 4*(days//1461); days %= 1461
        if days > 365: jy += (days-1)//365; days = (days-1)%365
        if days < 186: jm = 1+(days//31); jd = 1+(days%31)
        else: jm = 7+((days-186)//30); jd = 1+((days-186)%30)
        return jy, jm, jd
    @staticmethod
    def today_str():
        now = datetime.now()
        jy, jm, jd = PersianDate.gregorian_to_jalali(now.year, now.month, now.day)
        return f"{jy}/{jm:02d}/{jd:02d}"
    @staticmethod
    def now_jalali():
        return PersianDate.today_str()
    @staticmethod
    def jalali_to_gregorian(jy, jm, jd):
        """تبدیل تاریخ شمسی به میلادی"""
        jy += 1595; days = -355668 + (365*jy) + (jy//33)*8 + (jy%33+3)//4 + jd
        if jm < 7: days += (jm-1)*31
        else: days += (jm-7)*30 + 186
        gy = 400*(days//146097); days %= 146097
        gy += 100*(days//36524); days %= 36524
        gy += 4*(days//1461); days %= 1461
        gy += (days-1)//365; days %= 365
        if days < 187: gm = 1+(days//31); gd = 1+(days%31)
        else: gm = 7+((days-186)//30); gd = 1+((days-186)%30)
        return gy, gm, gd

# ─── Default Data ───
DEFAULT_EMPLOYEES = [
    {"name":"مریم","specialty":"مو","phone":"","share_percent":0,"salary":0,"pay_type":"ماهانه","start_date":"","status":"فعال","deductions":0,"deduction_note":""},
    {"name":"زهرا","specialty":"ناخن","phone":"","share_percent":0,"salary":0,"pay_type":"ماهانه","start_date":"","status":"فعال","deductions":0,"deduction_note":""},
    {"name":"سارا","specialty":"ابرو","phone":"","share_percent":0,"salary":0,"pay_type":"ماهانه","start_date":"","status":"فعال","deductions":0,"deduction_note":""},
    {"name":"نیلوفر","specialty":"مژه","phone":"","share_percent":0,"salary":0,"pay_type":"ماهانه","start_date":"","status":"فعال","deductions":0,"deduction_note":""},
]
DEFAULT_SERVICES = [
    {"name":"رنگ مو","category":"مو","default_price":80000},
    {"name":"مش","category":"مو","default_price":120000},
    {"name":"کوتاهی مو","category":"مو","default_price":50000},
    {"name":"شینیون","category":"مو","default_price":100000},
    {"name":"فر دائمی","category":"مو","default_price":150000},
    {"name":"ابرو برداشتن","category":"ابرو","default_price":20000},
    {"name":"ابرو رنگ","category":"ابرو","default_price":15000},
    {"name":"تاتو ابرو","category":"ابرو","default_price":80000},
    {"name":"کاشت مژه","category":"مژه","default_price":100000},
    {"name":"اکستنشن مژه","category":"مژه","default_price":150000},
    {"name":"جلاس مژه","category":"مژه","default_price":50000},
    {"name":"مانیکور","category":"ناخن","default_price":60000},
    {"name":"پدیکور","category":"ناخن","default_price":70000},
    {"name":"ژل ناخن","category":"ناخن","default_price":80000},
    {"name":"کاشت ناخن","category":"ناخن","default_price":120000},
    {"name":"فرچ ناخن","category":"ناخن","default_price":40000},
]
DEFAULT_EXPENSE_CATEGORIES = ["مواد مصرفی","اجاره","قبوض","حقوق ثابت","تبلیغات","لوازم","متفرقه"]

# ─── Excel Manager ───
PINK = PatternFill(start_color="D81B60", end_color="D81B60", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFF")

def _ensure_workbook(fp, headers):
    if not os.path.exists(fp):
        wb = Workbook(); ws = wb.active; ws.title = "داده‌ها"
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = HDR_FONT; cell.fill = PINK; cell.alignment = Alignment(horizontal="center")
        wb.save(fp)
    return load_workbook(fp)

def _read_rows(fp):
    if not os.path.exists(fp): return []
    wb = load_workbook(fp); ws = wb.active
    return [row for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]

def init_all():
    _ensure_workbook(EMPLOYEES_FILE, ["نام","تخصص","تلفن","درصد سهم","حقوق ثابت","نوع پرداخت","تاریخ شروع","وضعیت","کسورات","توضیحات کسورات"])
    ws = load_workbook(EMPLOYEES_FILE).active
    if ws.max_row <= 1:
        wb = load_workbook(EMPLOYEES_FILE); ws = wb.active
        for e in DEFAULT_EMPLOYEES: ws.append([e["name"],e["specialty"],e["phone"],e["share_percent"],e["salary"],e["pay_type"],e["start_date"],e["status"],e["deductions"],e["deduction_note"]])
        wb.save(EMPLOYEES_FILE)
    # Migration: ensure new employee columns exist (for pre-existing files)
    if os.path.exists(EMPLOYEES_FILE):
        wb = load_workbook(EMPLOYEES_FILE); ws = wb.active
        want = ["نام","تخصص","تلفن","درصد سهم","حقوق ثابت","نوع پرداخت","تاریخ شروع","وضعیت","کسورات","توضیحات کسورات"]
        if ws.max_column < 10:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=9, value=0)
                ws.cell(row=r, column=10, value="")
            for c, h in enumerate(want, 1):
                ws.cell(row=1, column=c, value=h).font = HDR_FONT
                ws.cell(row=1, column=c).fill = PINK
            wb.save(EMPLOYEES_FILE)
    _ensure_workbook(SERVICES_FILE, ["نام خدمت","دسته‌بندی","قیمت پیش‌فرض"])
    ws2 = load_workbook(SERVICES_FILE).active
    if ws2.max_row <= 1:
        wb2 = load_workbook(SERVICES_FILE); ws2 = wb2.active
        for s in DEFAULT_SERVICES: ws2.append([s["name"],s["category"],s["default_price"]])
        wb2.save(SERVICES_FILE)
    _ensure_workbook(TRANSACTIONS_FILE, [
        "تاریخ","نام مشتری","نام خدمت","دسته‌بندی","نام کارمند",
        "مبلغ خالص","تخفیف","مبلغ نهایی","روش پرداخت","پورسانت کارمند","سهم سالن","یادداشت","انعام"
    ])
    _ensure_workbook(CUSTOMERS_FILE, ["نام","تلفن","تخصص مورد علاقه","تاریخ عضویت","تاریخ تولد","تعداد مراجعه","یادداشت","امتیاز"])
    _ensure_workbook(EXPENSES_FILE, ["تاریخ","دسته‌بندی","مبلغ","توضیحات","روش پرداخت","یادداشت"])
    _ensure_workbook(INVENTORY_FILE, ["نام ماده","واحد","موجودی فعلی","حداقل موجودی","قیمت واحد (تومان)"])
    _ensure_workbook(INVENTORY_LOG_FILE, ["تاریخ","نوع","نام ماده","تعداد","علت/توضیح","لینک خدمت"])
    # Migration: ensure 'tip' column exists in transactions
    if os.path.exists(TRANSACTIONS_FILE):
        wb = load_workbook(TRANSACTIONS_FILE); ws = wb.active
        if ws.max_column < 13:
            ws.cell(row=1, column=13, value="انعام").font = HDR_FONT; ws.cell(row=1, column=13).fill = PINK
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=13).value is None:
                    ws.cell(row=r, column=13, value=0)
            wb.save(TRANSACTIONS_FILE)
    # Migration: ensure 'birth_date' column exists in customers (index 4)
    if os.path.exists(CUSTOMERS_FILE):
        wb = load_workbook(CUSTOMERS_FILE); ws = wb.active
        if ws.max_column < 8:
            # shift existing columns if needed: old layout had 7 cols (no birth_date at idx4)
            # old: name,phone,specialty,join_date,visit_count,note,points
            # new: name,phone,specialty,join_date,birth_date,visit_count,note,points
            for r in range(2, ws.max_row + 1):
                visit = ws.cell(row=r, column=5).value
                note = ws.cell(row=r, column=6).value
                points = ws.cell(row=r, column=7).value if ws.max_column >= 7 else None
                ws.cell(row=r, column=6, value=visit)
                ws.cell(row=r, column=7, value=note)
                ws.cell(row=r, column=8, value=points)
                ws.cell(row=r, column=5, value="")
            ws.cell(row=1, column=5, value="تاریخ تولد").font = HDR_FONT
            ws.cell(row=1, column=5).fill = PINK
            wb.save(CUSTOMERS_FILE)
    # NOTE: auth.init_auth() is called at module import time below, so it runs
    # under both `run.py` and Gunicorn (which imports web_app:app directly).

# ─── Data Access: Employees ───
def get_employees():
    rows = _read_rows(EMPLOYEES_FILE)
    result = []
    for r in rows:
        result.append({
            "name":str(r[0]),
            "specialty":str(r[1] or ""),
            "phone":str(r[2] or ""),
            "share_percent":float(r[3] or 0),
            "salary":int(r[4] or 0),
            "pay_type":str(r[5] or "ماهانه"),
            "start_date":str(r[6] or ""),
            "status":str(r[7] or "فعال"),
            "deductions":int(r[8] or 0),
            "deduction_note":str(r[9] or ""),
        })
    return result

def save_employees(emps):
    wb = Workbook(); ws = wb.active; ws.title = "داده‌ها"
    for c, h in enumerate(["نام","تخصص","تلفن","درصد سهم","حقوق ثابت","نوع پرداخت","تاریخ شروع","وضعیت","کسورات","توضیحات کسورات"], 1):
        cell = ws.cell(row=1, column=c, value=h); cell.font = HDR_FONT; cell.fill = PINK
    for e in emps:
        ws.append([e["name"],e["specialty"],e["phone"],e["share_percent"],e.get("salary",0),e.get("pay_type","ماهانه"),e.get("start_date",""),e.get("status","فعال"),e.get("deductions",0),e.get("deduction_note","")])
    wb.save(EMPLOYEES_FILE)

def employee_payroll(employees, monthly_txns, month=None):
    """محاسبه حقوق و دستمزد کارمندان برای تراکنش‌های یک ماه.
    month: e.g. "1404/05" — used to filter variable deductions
    returns: {emp_name: {salary, commission, ins_ded, var_ded, deductions, deduction_note, net, gross, pay_type, status}}
    """
    emp_shares = {e["name"]: e["share_percent"] for e in employees}
    comm_by_emp = defaultdict(int)
    for t in monthly_txns:
        c = t.get("commission", int(t["amount"]*emp_shares.get(t["employee"],0)/100))
        comm_by_emp[t["employee"]] += c
    var_ded = defaultdict(int)
    if month:
        for d in get_deductions(month):
            var_ded[d["employee"]] += d["amount"]
    result = {}
    for e in employees:
        name = e["name"]
        salary = e.get("salary", 0)
        commission = comm_by_emp.get(name, 0)
        ins_ded = e.get("deductions", 0)        # fixed insurance
        var = var_ded.get(name, 0)               # variable (loans, inventory, etc.)
        total_ded = ins_ded + var
        gross = salary + commission
        net = gross - total_ded
        result[name] = {
            "salary": salary,
            "commission": commission,
            "ins_ded": ins_ded,
            "var_ded": var,
            "deductions": total_ded,
            "deduction_note": e.get("deduction_note", ""),
            "gross": gross,
            "net": net,
            "pay_type": e.get("pay_type", "ماهانه"),
            "status": e.get("status", "فعال"),
        }
    return result

# ─── Data Access: Employee Deductions (variable, per month) ───
DED_FILE = os.path.join(DATA_DIR, "employee_deductions.xlsx")
DED_HEADERS = ["کارمند", "نوع", "مبلغ", "تاریخ", "ماه", "توضیحات"]

def get_deductions(month=None):
    rows = _read_rows(DED_FILE)
    result = []
    for r in rows:
        d = {"employee": str(r[0]), "type": str(r[1] or ""), "amount": int(r[2] or 0),
             "date": str(r[3] or ""), "month": str(r[4] or ""), "note": str(r[5] or "")}
        if month is None or d["month"] == month:
            result.append(d)
    return result

def add_deduction(employee, dtype, amount, date, month, note):
    add_expense(date, "کسورات کارمند", amount, f"کسورات {employee} — {dtype}: {note}", "نقدی", "ثبت از فیش حقوقی")
    rows = _read_rows(DED_FILE)
    rows.append([employee, dtype, amount, date, month, note])
    wb = Workbook(); ws = wb.active; ws.title = "کسورات"
    for c, h in enumerate(DED_HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h); cell.font = HDR_FONT; cell.fill = PINK
    for r in rows:
        ws.append(list(r))
    wb.save(DED_FILE)

def delete_deduction(idx):
    rows = _read_rows(DED_FILE)
    if 0 <= idx < len(rows):
        rows.pop(idx)
        wb = Workbook(); ws = wb.active; ws.title = "کسورات"
        for c, h in enumerate(DED_HEADERS, 1):
            cell = ws.cell(row=1, column=c, value=h); cell.font = HDR_FONT; cell.fill = PINK
        for r in rows:
            ws.append(list(r))
        wb.save(DED_FILE)

# ─── Data Access: Services ───
def get_services():
    rows = _read_rows(SERVICES_FILE)
    return [{"name":str(r[0]),"category":str(r[1] or ""),"default_price":int(r[2] or 0)} for r in rows]

def save_services(svcs):
    wb = Workbook(); ws = wb.active; ws.title = "داده‌ها"
    for c, h in enumerate(["نام خدمت","دسته‌بندی","قیمت پیش‌فرض"], 1):
        cell = ws.cell(row=1, column=c, value=h); cell.font = HDR_FONT; cell.fill = PINK
    for s in svcs: ws.append([s["name"],s["category"],s["default_price"]])
    wb.save(SERVICES_FILE)

def get_service_category(service_name):
    """پیدا کردن دسته‌بندی یک خدمت"""
    svcs = get_services()
    svc = next((s for s in svcs if s["name"] == service_name), None)
    return svc["category"] if svc else ""

# ─── Data Access: Customers ───
def get_customers():
    rows = _read_rows(CUSTOMERS_FILE)
    result = []
    for r in rows:
        cust = {"name":str(r[0]),"phone":str(r[1] or ""),"specialty":str(r[2] or ""),"join_date":str(r[3] or ""),
                "birth_date":str(r[4] or ""),"visit_count":int(r[5] or 0),"note":str(r[6] or "")}
        cust["points"] = int(r[7]) if len(r) > 7 and r[7] else 0
        result.append(cust)
    return result

def save_customers(custs):
    wb = Workbook(); ws = wb.active; ws.title = "داده‌ها"
    for c, h in enumerate(["نام","تلفن","تخصص مورد علاقه","تاریخ عضویت","تاریخ تولد","تعداد مراجعه","یادداشت","امتیاز"], 1):
        cell = ws.cell(row=1, column=c, value=h); cell.font = HDR_FONT; cell.fill = PINK
    for cu in custs: ws.append([cu["name"],cu["phone"],cu["specialty"],cu["join_date"],cu.get("birth_date",""),cu["visit_count"],cu["note"],cu.get("points",0)])
    wb.save(CUSTOMERS_FILE)

# ─── Data Access: Transactions ───
def get_transactions(start_date=None, end_date=None):
    if not os.path.exists(TRANSACTIONS_FILE): return []
    wb = load_workbook(TRANSACTIONS_FILE); ws = wb.active; txns = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        t = {
            "date":str(row[0]),"customer":str(row[1]),"service":str(row[2]),
            "category":str(row[3]) if row[3] else "",
            "employee":str(row[4]),
            "amount":int(row[5] or 0),
            "discount":int(row[6] or 0),
            "final_amount":int(row[7] or 0),
            "payment_method":str(row[8] or "نقدی"),
            "commission":int(row[9] or 0),
            "salon_share":int(row[10] or 0),
            "note":str(row[11] or ""),
            "tip":int(row[12] or 0) if len(row) > 12 else 0
        }
        if start_date and end_date:
            if start_date <= t["date"] <= end_date: txns.append(t)
        else: txns.append(t)
    return txns

def add_transaction(date_str, customer, service, category, employee, amount, discount=0, final_amount=0, payment_method="نقدی", commission=0, salon_share=0, note="", tip=0):
    wb = load_workbook(TRANSACTIONS_FILE); ws = wb.active
    ws.append([date_str, customer, service, category, employee, int(amount), int(discount), int(final_amount), payment_method, int(commission), int(salon_share), note, int(tip)])
    wb.save(TRANSACTIONS_FILE)

def delete_transaction(index):
    wb = load_workbook(TRANSACTIONS_FILE); ws = wb.active
    ws.delete_rows(index + 2); wb.save(TRANSACTIONS_FILE)

# ─── Data Access: Expenses ───
def get_expenses(start_date=None, end_date=None):
    if not os.path.exists(EXPENSES_FILE): return []
    wb = load_workbook(EXPENSES_FILE); ws = wb.active; expenses = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        e = {
            "date":str(row[0]),"category":str(row[1]),
            "amount":int(row[2] or 0),"description":str(row[3] or ""),
            "payment_method":str(row[4] or "نقدی"),"note":str(row[5] or "")
        }
        if start_date and end_date:
            if start_date <= e["date"] <= end_date: expenses.append(e)
        else: expenses.append(e)
    return expenses

def add_expense(date_str, category, amount, description="", payment_method="نقدی", note=""):
    wb = load_workbook(EXPENSES_FILE); ws = wb.active
    ws.append([date_str, category, int(amount), description, payment_method, note])
    wb.save(EXPENSES_FILE)

def delete_expense(index):
    wb = load_workbook(EXPENSES_FILE); ws = wb.active
    ws.delete_rows(index + 2); wb.save(EXPENSES_FILE)

# ─── Data Access: Inventory ───
def get_inventory():
    if not os.path.exists(INVENTORY_FILE): return []
    rows = _read_rows(INVENTORY_FILE)
    return [{"name":str(r[0]),"unit":str(r[1] or "عدد"),"stock":int(r[2] or 0),"min_stock":int(r[3] or 0),"unit_price":int(r[4] or 0)} for r in rows]

def save_inventory(items):
    wb = Workbook(); ws = wb.active; ws.title = "انبار"
    for c, h in enumerate(["نام ماده","واحد","موجودی فعلی","حداقل موجودی","قیمت واحد (تومان)"], 1):
        cell = ws.cell(row=1, column=c, value=h); cell.font = HDR_FONT; cell.fill = PINK
    for it in items: ws.append([it["name"],it["unit"],it["stock"],it["min_stock"],it["unit_price"]])
    wb.save(INVENTORY_FILE)

def get_inventory_log(item_name=None):
    if not os.path.exists(INVENTORY_LOG_FILE): return []
    wb = load_workbook(INVENTORY_LOG_FILE); ws = wb.active; logs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        log = {"date":str(row[0]),"type":str(row[1]),"item":str(row[2]),"qty":int(row[3] or 0),"reason":str(row[4] or ""),"service_link":str(row[5] or "")}
        if item_name:
            if log["item"] == item_name: logs.append(log)
        else: logs.append(log)
    return logs

def add_inventory_log(date_str, log_type, item, qty, reason="", service_link=""):
    wb = load_workbook(INVENTORY_LOG_FILE); ws = wb.active
    ws.append([date_str, log_type, item, int(qty), reason, service_link])
    wb.save(INVENTORY_LOG_FILE)

# ─── Backup Helper ───
def create_backup_excel(transactions, filename):
    wb = Workbook(); ws = wb.active; ws.title = "تراکنش‌ها"
    headers = ["تاریخ","مشتری","خدمت","دسته","کارمند","مبلغ خالص","تخفیف","مبلغ نهایی","روش پرداخت","پورسانت","سهم سالن","یادداشت","انعام"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h); cell.font = HDR_FONT; cell.fill = PINK
    for r, t in enumerate(transactions, 2):
        ws.cell(row=r,column=1,value=t["date"]); ws.cell(row=r,column=2,value=t["customer"])
        ws.cell(row=r,column=3,value=t["service"]); ws.cell(row=r,column=4,value=t.get("category",""))
        ws.cell(row=r,column=5,value=t["employee"]); ws.cell(row=r,column=6,value=t["amount"])
        ws.cell(row=r,column=7,value=t.get("discount",0)); ws.cell(row=r,column=8,value=t.get("final_amount",t["amount"]))
        ws.cell(row=r,column=9,value=t.get("payment_method","نقدی")); ws.cell(row=r,column=10,value=t.get("commission",0))
        ws.cell(row=r,column=11,value=t.get("salon_share",0)); ws.cell(row=r,column=12,value=t.get("note",""))
        ws.cell(row=r,column=13,value=t.get("tip",0))
    for c in range(1,14): ws.column_dimensions[get_column_letter(c)].width = 16
    fp = os.path.join(DATA_DIR, filename); wb.save(fp); return fp

def _file_size(fp):
    s = os.path.getsize(fp)
    return f"{s//1024} KB" if s < 1048576 else f"{s//1048576} MB"

# ─── Auth Routes ───
@app.route("/login", methods=["GET", "POST"])
@csrf_required
def login():
    if request.method == "POST":
        ip = request.remote_addr
        allowed, retry = check_rate_limit(ip)
        if not allowed:
            flash(f"تلاش‌های ناموفق زیاد بود. لطفاً {retry} ثانیه صبر کنید", "error")
            return render_template("login.html")
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = auth.authenticate(username, password)
        if user:
            auth.login_user(user)
            reset_fails(ip)
            flash(f"خوش‌آمدید {user['name']} ({auth.ROLES.get(user['role'], user['role'])}", "success")
            # Redirect to the first page the role is allowed to access (avoid loop for reception/employee)
            first_allowed = next((r for r in ["/", "/customers", "/dashboard", "/reports"] if auth.role_can_access(user["role"], r)), "/login")
            return redirect(first_allowed)
        register_fail(ip)
        flash("نام کاربری یا رمز عبور اشتباه است", "error")
    # If already logged in, go to dashboard
    if "user" in session:
        return redirect(url_for("dashboard"))
    return render_template("login.html")

@app.route("/logout")
def logout():
    auth.logout_user()
    flash("با موفقیت خارج شدید", "info")
    return redirect(url_for("login"))

@app.route("/change_password", methods=["GET", "POST"])
@login_required
@csrf_required
def change_password():
    if request.method == "POST":
        current = request.form.get("current_password", "")
        new_pw = request.form.get("new_password", "")
        confirm = request.form.get("confirm_password", "")
        user = auth.get_user(session["user"])
        # Validate current password
        if not auth.check_password(current, user["password_hash"]):
            flash("رمز عبور فعلی اشتباه است", "error")
        # Validate new password policy
        elif not re.fullmatch(r"^(?=.*[A-Za-z])(?=.*\d).{8,}$", new_pw):
            flash("رمز جدید باید حداقل ۸ کاراکتر و شامل حرف و عدد باشد", "error")
        elif new_pw != confirm:
            flash("تکرار رمز جدید مطابقت ندارد", "error")
        else:
            auth.update_password(user["username"], new_pw)
            flash("رمز عبور با موفقیت تغییر یافت", "success")
            return redirect(url_for("dashboard"))
    return render_template("change_password.html")

@app.route("/users", methods=["GET", "POST"])
@auth.role_required("admin")
def manage_users():
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            uname = request.form.get("username", "").strip()
            pw = request.form.get("password", "")
            name = request.form.get("name", "").strip()
            role = request.form.get("role", "employee")
            if not uname or not pw:
                flash("نام کاربری و رمز عبور الزامی است", "error")
            elif auth.get_user(uname):
                flash("این نام کاربری قبلاً ثبت شده", "error")
            else:
                auth.create_user(uname, pw, name, role)
                flash(f"کاربر «{uname}» با نقش {auth.ROLES.get(role, role)} ایجاد شد", "success")
        elif action == "delete":
            uname = request.form.get("username", "")
            if uname == "admin":
                flash("نمی‌توان کاربر مدیر اصلی را حذف کرد", "error")
            elif uname == session.get("user"):
                flash("شما نمی‌توانید حساب کاربری خود را حذف کنید", "error")
            else:
                conn = auth.get_db(); conn.execute("DELETE FROM users WHERE username=?", (uname,)); conn.commit(); conn.close()
                flash(f"کاربر «{uname}» حذف شد", "success")
        return redirect(url_for("manage_users"))
    conn = auth.get_db()
    users = [dict(r) for r in conn.execute("SELECT id, username, name, role, active, created_at FROM users ORDER BY id").fetchall()]
    conn.close()
    return render_template("users.html", users=users, roles=auth.ROLES)

# ─── Dashboard ───
@app.route("/dashboard")
def dashboard():
    today = PersianDate.today_str()
    all_txns = get_transactions()
    today_txns = [t for t in all_txns if t["date"] == today]

    # Today stats
    today_total = sum(t["final_amount"] for t in today_txns)
    today_customers = len(set(t["customer"] for t in today_txns))
    today_count = len(today_txns)
    today_commission = sum(t["commission"] for t in today_txns)

    # This month
    now = datetime.now()
    jy, jm, jd = PersianDate.gregorian_to_jalali(now.year, now.month, now.day)
    month_start = f"{jy}/{jm:02d}/01"
    month_end = f"{jy}/{jm+1:02d}/01" if jm < 12 else f"{jy+1}/01/01"
    month_txns = [t for t in all_txns if month_start <= t["date"] < month_end]
    month_total = sum(t["final_amount"] for t in month_txns)
    month_expenses = sum(e["amount"] for e in get_expenses(start_date=month_start, end_date=month_end.replace("/01","/31")))
    month_profit = month_total - month_expenses

    # Total fixed salary for current month (all active employees)
    _emps = get_employees()
    month_salary_total = sum(e["salary"] for e in _emps if e.get("status", "فعال") == "فعال")

    # Last 7 days chart data
    chart_labels = []
    chart_income = []
    chart_expenses_list = []
    for i in range(6, -1, -1):
        d = datetime.now() - timedelta(days=i)
        jy2, jm2, jd2 = PersianDate.gregorian_to_jalali(d.year, d.month, d.day)
        date_str = f"{jy2}/{jm2:02d}/{jd2:02d}"
        chart_labels.append(f"{jd2}/{jm2}")
        day_txns = [t for t in all_txns if t["date"] == date_str]
        day_exp = [e for e in get_expenses(start_date=date_str, end_date=date_str)]
        chart_income.append(sum(t["final_amount"] for t in day_txns))
        chart_expenses_list.append(sum(e["amount"] for e in day_exp))

    # Low stock alerts
    inventory = get_inventory()
    low_stock = [it for it in inventory if it["stock"] <= it["min_stock"]]

    # Birthday alerts: customers whose birth date (MM/DD) is today OR in 2 days
    today_md = today[5:]  # "MM/DD" portion of "YYYY/MM/DD"
    # compute Jalali date for 2 days from now
    d2 = datetime.now() + timedelta(days=2)
    jy2, jm2, jd2 = PersianDate.gregorian_to_jalali(d2.year, d2.month, d2.day)
    in2_md = f"{jm2:02d}/{jd2:02d}"
    birthday_customers = []
    birthday_soon = []  # 2 days before birthday
    for c in get_customers():
        bd = c.get("birth_date", "")
        if bd and len(bd) >= 5:
            if bd[5:] == today_md:
                birthday_customers.append(c["name"])
            elif bd[5:] == in2_md:
                birthday_soon.append(c["name"])

    # Service category breakdown for pie chart
    cat_totals = defaultdict(int)
    for t in month_txns:
        cat = t.get("category", "") or "نامشخص"
        cat_totals[cat] += t["final_amount"]

    return render_template("dashboard.html",
        today=today, today_total=today_total, today_customers=today_customers,
        today_count=today_count, today_commission=today_commission,
        month_total=month_total, month_expenses=month_expenses, month_profit=month_profit,
        chart_labels=json.dumps(chart_labels, ensure_ascii=False),
        chart_income=json.dumps(chart_income),
        chart_expenses=json.dumps(chart_expenses_list),
        low_stock=low_stock,
        cat_labels=json.dumps(list(cat_totals.keys()), ensure_ascii=False),
        cat_values=json.dumps(list(cat_totals.values())),
        birthday_customers=birthday_customers,
        birthday_soon=birthday_soon,
        month_salary_total=month_salary_total
    )

# ─── Routes: Index ───
@app.route("/")
@login_required
def index():
    return render_template("index.html", today=PersianDate.today_str(),
        services=get_services(), employees=get_employees())

@login_required
@app.route("/submit_transaction", methods=["POST"])
def submit_transaction():
    data = request.form
    customer = data.get("customer_name","").strip() or data.get("customer_name_manual","").strip()
    phone = data.get("customer_phone","").strip()
    services = data.getlist("service[]")
    employees = data.getlist("employee[]")
    amounts = data.getlist("amount[]")
    discount_type = data.get("discount_type", "amount")
    discount_value = int(data.get("discount_value", 0) or 0)
    payment_method = data.get("payment_method", "نقدی")
    tip = int(data.get("tip", 0) or 0)
    deposit = int(data.get("deposit", 0) or 0)
    deposit_date = data.get("deposit_date", "").strip()
    deposit_method = data.get("deposit_method", "").strip()
    deposit_note = data.get("deposit_note", "").strip()
    split_methods = data.getlist("split_method[]")
    split_amounts = data.getlist("split_amount[]")
    if split_methods and split_amounts:
        payment_summary = ", ".join([f"{m}: {int(a):,}" for m, a in zip(split_methods, split_amounts) if a])
    else:
        payment_summary = payment_method
    send_sms = data.get("send_sms", "")
    use_points = data.get("use_points", "")

    if not customer:
        flash("لطفاً نام مشتری را وارد کنید", "error"); return redirect(url_for("index"))
    if not services:
        flash("حداقل یک خدمت اضافه کنید", "error"); return redirect(url_for("index"))

    today = PersianDate.today_str()
    emps = get_employees()
    emp_shares = {e["name"]: e["share_percent"] for e in emps}

    # Calculate subtotal
    subtotal = 0
    items_data = []
    for svc, emp, amt in zip(services, employees, amounts):
        if not amt or not amt.isdigit(): continue
        amount = int(amt)
        subtotal += amount
        emp_name = emp.split(" (")[0] if " (" in emp else emp
        share = emp_shares.get(emp_name, 0)
        commission = int(amount * share / 100)
        salon_share = amount - commission
        # FIX: look up service category
        category = get_service_category(svc.split(" - ")[0])
        items_data.append({
            "service": svc.split(" - ")[0],
            "employee": emp_name,
            "amount": amount,
            "commission": commission,
            "salon_share": salon_share,
            "category": category
        })

    # Calculate discount
    if discount_type == "percent":
        discount_amount = int(subtotal * discount_value / 100)
    else:
        discount_amount = discount_value

    final_amount = max(0, subtotal - discount_amount)
    # Deposit already paid earlier (reservation) -> subtract from final payable
    payable_amount = max(0, final_amount - deposit)

    # Use points if selected
    points_used = 0
    if use_points == "on" and customer:
        custs = get_customers()
        cust = next((c for c in custs if c["name"] == customer), None)
        if cust and cust.get("points", 0) > 0:
            points_used = min(cust["points"], final_amount // 1000)
            final_amount = max(0, final_amount - points_used * 1000)
            cust["points"] -= points_used
            save_customers(custs)

    # Add transactions (FIX: pass category, tip only on first item)
    for idx, item in enumerate(items_data):
        item_tip = tip if idx == 0 else 0
        add_transaction(
            today, customer, item["service"], item["category"], item["employee"],
            item["amount"], discount_amount // len(items_data) if items_data else 0,
            final_amount // len(items_data) if items_data else 0,
            payment_method, item["commission"], item["salon_share"], phone, item_tip
        )

    # Update customer visit count and points
    if customer:
        custs = get_customers()
        cust = next((c for c in custs if c["name"] == customer), None)
        if cust:
            cust["visit_count"] += 1
            cust["points"] = cust.get("points", 0) + (final_amount // 10000)
            save_customers(custs)
        else:
            custs.append({
                "name": customer, "phone": phone, "specialty": "",
                "join_date": today, "visit_count": 1, "note": "",
                "points": final_amount // 10000
            })
            save_customers(custs)

    # Auto-decrement inventory for used services
    inventory = get_inventory()
    if inventory:
        svc_names = [it["service"] for it in items_data]
        # Simple: if service name matches inventory item name, decrement by 1
        for inv in inventory:
            if inv["name"] in svc_names:
                used_qty = svc_names.count(inv["name"])
                inv["stock"] = max(0, inv["stock"] - used_qty)
                add_inventory_log(today, "خروج", inv["name"], used_qty, "مصرف خودکار", ", ".join(set(svc_names)))
        save_inventory(inventory)

    if payment_method == "نسیه":
        flash(f"⚠️ فاکتور نسیه ثبت شد! مشتری: {customer} — مبلغ کل: {final_amount:,} تومان", "info")
    else:
        if deposit > 0:
            flash(f"✅ تراکنش ثبت شد! مشتری: {customer} — مبلغ کل: {final_amount:,} — بیعانه قبلی: {deposit:,} — مانده قابل پرداخت: {payable_amount:,} تومان — پرداخت: {payment_summary}", "success")
        else:
            flash(f"✅ تراکنش ثبت شد! مشتری: {customer} — مبلغ نهایی: {final_amount:,} تومان — پرداخت: {payment_summary}", "success")

    if tip > 0:
        flash(f"💰 انعام ثبت شد: {tip:,} تومان", "success")

    if send_sms == "on":
        flash(f"📱 پیامک فاکتور برای {customer} ارسال خواهد شد (قابلیت در دست ساخت)", "info")

    # ─── Generate PDF invoice ───
    invoice_no = f"INV-{today.replace('/','')}-{int(datetime.now().timestamp())%100000:05d}"
    pdf_dir = os.path.join(DATA_DIR, "invoices")
    os.makedirs(pdf_dir, exist_ok=True)
    pdf_path = os.path.join(pdf_dir, f"{invoice_no}.pdf")
    try:
        emp_commissions = ", ".join([f"{it['employee']}: {it['commission']:,}" for it in items_data])
        invoice_pdf.generate_invoice({
            "salon_name": "Heli Beauty Studio",
            "invoice_no": invoice_no,
            "date": today,
            "customer": customer,
            "phone": phone,
            "items": items_data,
            "subtotal": subtotal,
            "discount": discount_amount,
            "tip": tip,
            "deposit": deposit,
            "deposit_date": deposit_date,
            "deposit_method": deposit_method,
            "deposit_note": deposit_note,
            "payment_method": payment_summary,
            "final_amount": final_amount,
            "payable_amount": payable_amount,
            "employee_commissions": emp_commissions,
        }, pdf_path)
        session["last_invoice"] = invoice_no
    except Exception as e:
        app.logger.error(f"PDF gen failed: {e}")

    return redirect(url_for("index"))

# ─── Routes: Reports ───
@app.route("/reports")
@login_required
def reports():
    today = PersianDate.today_str()
    all_txns = get_transactions()
    emp_filter = request.args.get("employee", "").strip()
    if emp_filter:
        txns = [t for t in all_txns if t["date"] == today and t.get("employee","") == emp_filter]
    else:
        txns = [t for t in all_txns if t["date"] == today]
    emps = get_employees()
    emp_shares = {e["name"]: e["share_percent"] for e in emps}
    total = sum(t.get("final_amount", t["amount"]) for t in txns)
    customers_count = len(set(t["customer"] for t in txns))
    tips = sum(t.get("tip", 0) for t in txns)
    # attach original index (in full transactions list) for edit/delete
    all_list = get_transactions()
    indexed = []
    for t in reversed(txns):
        # find its index in the full list (last match by identity of dict)
        idx = all_list.index(t) if t in all_list else -1
        indexed.append((idx, t))
    return render_template("reports.html", transactions=indexed,
        total=total, customers=customers_count, services_count=len(txns),
        tips=tips, today=today, employees=emps, selected_emp=emp_filter)

# ─── Routes: Transaction delete / edit ───
@app.route("/transaction/delete", methods=["POST"])
@login_required
@csrf_required
def transaction_delete():
    idx = int(request.form.get("idx", -1))
    if idx >= 0:
        delete_transaction(idx)
        flash("🗑️ فاکتور حذف شد", "info")
    return redirect(request.referrer or url_for("reports"))

@app.route("/transaction/edit", methods=["GET","POST"])
@login_required
@csrf_required
def transaction_edit():
    idx = int(request.args.get("idx", request.form.get("idx", -1)))
    txns = get_transactions()
    if idx < 0 or idx >= len(txns):
        flash("فاکتور مورد نظر یافت نشد", "error")
        return redirect(url_for("reports"))
    t = txns[idx]
    if request.method == "POST":
        action = request.form.get("action")
        if action == "save":
            t["customer"] = request.form.get("customer", t["customer"])
            t["service"] = request.form.get("service", t["service"])
            t["employee"] = request.form.get("employee", t["employee"])
            t["amount"] = int(request.form.get("amount", 0) or 0)
            t["discount"] = int(request.form.get("discount", 0) or 0)
            t["final_amount"] = int(request.form.get("final_amount", 0) or 0)
            t["payment_method"] = request.form.get("payment_method", "نقدی")
            t["note"] = request.form.get("note", "")
            txns[idx] = t
            _save_all_transactions(txns)
            flash("✅ فاکتور ویرایش شد", "success")
            return redirect(url_for("reports"))
    emps = get_employees()
    return render_template("transaction_edit.html", t=t, idx=idx, employees=emps)

def _save_all_transactions(txns):
    wb = Workbook(); ws = wb.active; ws.title = "داده‌ها"
    for c, h in enumerate(["تاریخ","نام مشتری","نام خدمت","دسته‌بندی","نام کارمند",
        "مبلغ خالص","تخفیف","مبلغ نهایی","روش پرداخت","پورسانت کارمند","سهم سالن","یادداشت","انعام"], 1):
        cell = ws.cell(row=1, column=c, value=h); cell.font = HDR_FONT; cell.fill = PINK
    for t in txns:
        ws.append([t.get("date",""),t.get("customer",""),t.get("service",""),t.get("category",""),
            t.get("employee",""),t.get("amount",0),t.get("discount",0),t.get("final_amount",t.get("amount",0)),
            t.get("payment_method","نقدی"),t.get("commission",0),t.get("salon_share",0),t.get("note",""),t.get("tip",0)])
    wb.save(TRANSACTIONS_FILE)

# ─── Routes: Monthly ───
@app.route("/monthly", methods=["GET","POST"])
@login_required
def monthly():
    now = datetime.now()
    jy, jm, jd = PersianDate.gregorian_to_jalali(now.year, now.month, now.day)
    month_str = request.args.get("month", f"{jy}/{jm:02d}")
    try:
        parts = month_str.split("/"); y, m = int(parts[0]), int(parts[1])
    except: y, m = jy, jm; month_str = f"{y}/{m:02d}"
    start = f"{y}/{m:02d}/01"
    end = f"{y+1}/01/01" if m == 12 else f"{y}/{m+1:02d}/01"
    all_txns = get_transactions()
    monthly_txns = [t for t in all_txns if start <= t["date"] < end]
    emp_shares = {e["name"]: e["share_percent"] for e in get_employees()}
    emp_totals = defaultdict(lambda: {"count":0,"amount":0,"commission":0})
    for t in monthly_txns:
        c = t.get("commission", int(t["amount"]*emp_shares.get(t["employee"],0)/100))
        emp_totals[t["employee"]]["count"] += 1
        emp_totals[t["employee"]]["amount"] += t.get("final_amount", t["amount"])
        emp_totals[t["employee"]]["commission"] += c
    emp_cat = defaultdict(lambda: defaultdict(lambda: {"count":0,"amount":0,"commission":0}))
    for t in monthly_txns:
        c = t.get("commission", int(t["amount"]*emp_shares.get(t["employee"],0)/100))
        cat = t.get("category", "") or "نامشخص"
        emp_cat[t["employee"]][cat]["count"] += 1
        emp_cat[t["employee"]][cat]["amount"] += t.get("final_amount", t["amount"])
        emp_cat[t["employee"]][cat]["commission"] += c
    details = []
    for en in sorted(emp_cat.keys()):
        for cn in sorted(emp_cat[en].keys()):
            d = emp_cat[en][cn]
            details.append({"employee":en,"category":cn,"count":d["count"],"amount":d["amount"],"commission":d["commission"]})
    grand_total = sum(t.get("final_amount", t["amount"]) for t in monthly_txns)
    grand_comm = sum(t.get("commission", int(t["amount"]*emp_shares.get(t["employee"],0)/100)) for t in monthly_txns)
    grand_tips = sum(t.get("tip", 0) for t in monthly_txns)
    months = [f"{yy}/{mm:02d}" for yy in range(1400,1410) for mm in range(1,13)]

    # Chart data: daily income for the month
    chart_labels = []
    chart_income = []
    days_in_month = 31
    for d in range(1, days_in_month + 1):
        ds = f"{y}/{m:02d}/{d:02d}"
        day_txns = [t for t in monthly_txns if t["date"] == ds]
        if day_txns or d <= 31:
            chart_labels.append(str(d))
            chart_income.append(sum(t.get("final_amount", t["amount"]) for t in day_txns))

    # Category breakdown for pie chart
    cat_totals = defaultdict(int)
    for t in monthly_txns:
        cat = t.get("category", "") or "نامشخص"
        cat_totals[cat] += t.get("final_amount", t["amount"])

    # Expenses for this month
    month_expenses = get_expenses(start_date=start, end_date=end)
    total_expenses = sum(e["amount"] for e in month_expenses)
    expense_cats = defaultdict(int)
    for e in month_expenses:
        expense_cats[e["category"]] += e["amount"]

    # Group transactions by employee for detailed view
    employee_transactions = defaultdict(list)
    for t in monthly_txns:
        employee_transactions[t["employee"]].append(t)

    # ─── Previous month comparison ───
    prev_m = m - 1
    prev_y = y
    if prev_m == 0:
        prev_m = 12; prev_y = y - 1
    prev_summary = _month_summary(prev_y, prev_m)
    prev_month_str = f"{prev_y}/{prev_m:02d}"
    def _delta(cur, prev):
        if prev > 0:
            return round((cur - prev) / prev * 100, 1)
        return None

    # Payroll for this month
    payroll = employee_payroll(get_employees(), monthly_txns)
    paid_salary = sum(e["amount"] for e in get_expenses(start_date=start, end_date=end) if e["category"] == "حقوق و دستمزد")
    total_salary = sum(p["salary"] for p in payroll.values())
    total_payroll = sum(p["net"] for p in payroll.values())

    return render_template("monthly.html", month_str=month_str, months=months,
        emp_totals=dict(emp_totals), details=details,
        grand_total=grand_total, grand_comm=grand_comm, grand_tips=grand_tips,
        total_count=len(monthly_txns),
        chart_labels=json.dumps(chart_labels, ensure_ascii=False),
        chart_income=json.dumps(chart_income),
        cat_labels=json.dumps(list(cat_totals.keys()), ensure_ascii=False),
        cat_values=json.dumps(list(cat_totals.values())),
        total_expenses=total_expenses,
        profit=grand_total - total_expenses,
        expense_cats=dict(expense_cats),
        employee_transactions=dict(employee_transactions),
        prev_month_str=prev_month_str,
        prev_summary=prev_summary,
        prev_income=prev_summary["income"],
        prev_expenses=prev_summary["expenses"],
        prev_profit=prev_summary["profit"],
        delta_income=_delta(grand_total, prev_summary["income"]),
        delta_expenses=_delta(total_expenses, prev_summary["expenses"]),
        delta_profit=_delta(grand_total - total_expenses, prev_summary["profit"]),
        payroll=payroll, total_salary=total_salary, total_payroll=total_payroll,
        paid_salary=paid_salary)

def _month_summary(y, m):
    """Compute income/expenses/profit for a given Jalali year/month."""
    start = f"{y}/{m:02d}/01"
    end = f"{y+1}/01/01" if m == 12 else f"{y}/{m+1:02d}/01"
    txns = [t for t in get_transactions() if start <= t["date"] < end]
    inc = sum(t.get("final_amount", t["amount"]) for t in txns)
    exp = sum(e["amount"] for e in get_expenses(start_date=start, end_date=end))
    tips = sum(t.get("tip", 0) for t in txns)
    comm = sum(t.get("commission", 0) for t in txns)
    return {"income": inc, "expenses": exp, "profit": inc - exp, "tips": tips, "commission": comm, "count": len(txns)}

@app.route("/monthly/export/<int:y>/<int:m>")
def export_monthly(y, m):
    month_str = f"{y}/{m:02d}"
    start = f"{y}/{m:02d}/01"
    end = f"{y+1}/01/01" if m == 12 else f"{y}/{m+1:02d}/01"
    monthly_txns = [t for t in get_transactions() if start <= t["date"] < end]
    if not monthly_txns: flash("تراکنشی وجود ندارد","error"); return redirect(url_for("monthly"))
    emp_shares = {e["name"]: e["share_percent"] for e in get_employees()}
    wb = Workbook(); ws = wb.active; ws.title = f"گزارش {y}-{m:02d}"
    headers = ["تاریخ","مشتری","خدمت","دسته","کارمند","مبلغ","پورسانت","انعام","یادداشت"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h); cell.font = HDR_FONT; cell.fill = PINK
    for r, t in enumerate(monthly_txns, 2):
        ws.cell(row=r,column=1,value=t["date"]); ws.cell(row=r,column=2,value=t["customer"])
        ws.cell(row=r,column=3,value=t["service"]); ws.cell(row=r,column=4,value=t.get("category",""))
        ws.cell(row=r,column=5,value=t["employee"]); ws.cell(row=r,column=6,value=t.get("final_amount",t["amount"]))
        ws.cell(row=r,column=7,value=t.get("commission",int(t["amount"]*emp_shares.get(t["employee"],0)/100)))
        ws.cell(row=r,column=8,value=t.get("tip",0)); ws.cell(row=r,column=9,value=t.get("note",""))
    for c in range(1,10): ws.column_dimensions[get_column_letter(c)].width = 18
    fp = os.path.join(DATA_DIR, f"report_{y}_{m:02d}.xlsx"); wb.save(fp)
    return send_file(fp, as_attachment=True)

# ─── Routes: Expenses ───
@app.route("/expenses", methods=["GET","POST"])
@login_required
def expenses_page():
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            date = request.form.get("date", PersianDate.today_str())
            cat = request.form.get("category", "")
            amount = int(request.form.get("amount", 0) or 0)
            desc = request.form.get("description", "")
            pm = request.form.get("payment_method", "نقدی")
            note = request.form.get("note", "")
            if amount > 0:
                add_expense(date, cat, amount, desc, pm, note)
                flash(f"✅ هزینه ثبت شد: {amount:,} تومان — {cat}", "success")
            else:
                flash("لطفاً مبلغ را وارد کنید", "error")
        elif action == "delete":
            idx = int(request.form.get("idx", 0))
            delete_expense(idx)
            flash("🗑️ هزینه حذف شد", "info")
        return redirect(url_for("expenses_page"))

    # Filter
    start_date = request.args.get("from", "")
    end_date = request.args.get("to", "")
    expenses = get_expenses(start_date=start_date if start_date else None, end_date=end_date if end_date else None)
    total = sum(e["amount"] for e in expenses)
    cat_totals = defaultdict(int)
    for e in expenses: cat_totals[e["category"]] += e["amount"]
    return render_template("expenses.html", expenses=list(reversed(expenses)),
        total=total, cat_totals=dict(cat_totals), today=PersianDate.today_str(),
        categories=DEFAULT_EXPENSE_CATEGORIES,
        start_date=start_date, end_date=end_date)

# ─── Routes: Inventory ───
@app.route("/inventory", methods=["GET","POST"])
@login_required
def inventory_page():
    if request.method == "POST":
        action = request.form.get("action")
        if action in ("add", "add_item"):
            name = request.form.get("name", "").strip()
            unit = request.form.get("unit", "عدد")
            stock = int(request.form.get("initial_stock", 0) or 0)
            min_stock = int(request.form.get("min_stock", 0) or 0)
            unit_price = int(request.form.get("unit_price", 0) or 0)
            if name:
                items = get_inventory()
                items.append({"name":name,"unit":unit,"stock":stock,"min_stock":min_stock,"unit_price":unit_price})
                save_inventory(items)
                add_inventory_log(PersianDate.today_str(), "ورود", name, stock, "ورود اولیه")
                flash(f"✅ ماده '{name}' به انبار اضافه شد", "success")
            else:
                flash("نام ماده را وارد کنید", "error")
        elif action == "stock_in":
            item_name = request.form.get("item_name", "")
            qty = int(request.form.get("qty", 0) or 0)
            reason = request.form.get("reason", "")
            items = get_inventory()
            item = next((i for i in items if i["name"] == item_name), None)
            if item and qty > 0:
                item["stock"] += qty
                save_inventory(items)
                add_inventory_log(PersianDate.today_str(), "ورود", item_name, qty, reason)
                flash(f"📥 ورود {qty} {item['unit']} {item_name}", "success")
        elif action == "stock_out":
            item_name = request.form.get("item_name", "")
            qty = int(request.form.get("qty", 0) or 0)
            reason = request.form.get("reason", "")
            items = get_inventory()
            item = next((i for i in items if i["name"] == item_name), None)
            if item and qty > 0:
                if item["stock"] >= qty:
                    item["stock"] -= qty
                    save_inventory(items)
                    add_inventory_log(PersianDate.today_str(), "خروج", item_name, qty, reason)
                    flash(f"📤 خروج {qty} {item['unit']} {item_name}", "success")
                else:
                    flash(f"موجودی کافی نیست! (موجود: {item['stock']})", "error")
        elif action == "delete_item":
            idx = int(request.form.get("idx", 0))
            items = get_inventory()
            items.pop(idx)
            save_inventory(items)
            flash("🗑️ ماده حذف شد", "info")
        return redirect(url_for("inventory_page"))

    inventory = get_inventory()
    alerts = [it for it in inventory if it["stock"] <= it["min_stock"]]
    total_value = sum(it["stock"] * it["unit_price"] for it in inventory)
    return render_template("inventory.html", inventory=inventory, alerts=alerts, total_value=total_value)

@app.route("/inventory/log/<item_name>")
@login_required
def inventory_log_page(item_name):
    logs = get_inventory_log(item_name)
    return render_template("inventory_log.html", item_name=item_name, logs=list(reversed(logs)))

# ─── Routes: Profit & Loss ───
@app.route("/profit_loss")
@login_required
def profit_loss():
    now = datetime.now()
    # Default: current month
    period = request.args.get("period", "month")
    start_date = request.args.get("from", "")
    end_date = request.args.get("to", "")

    if period == "today":
        start_date = end_date = PersianDate.today_str()
        title = f"سود و زیان امروز — {start_date}"
    elif period == "month":
        jy, jm, jd = PersianDate.gregorian_to_jalali(now.year, now.month, now.day)
        start_date = f"{jy}/{jm:02d}/01"
        end_date = f"{jy}/{jm+1:02d}/01" if jm < 12 else f"{jy+1}/01/01"
        title = f"سود و زیان ماه {jy}/{jm:02d}"
    elif period == "year":
        jy, jm, jd = PersianDate.gregorian_to_jalali(now.year, now.month, now.day)
        start_date = f"{jy}/01/01"
        end_date = f"{jy+1}/01/01"
        title = f"سود و زیان سال {jy}"
    elif period == "custom" and start_date and end_date:
        title = f"سود و زیان {start_date} تا {end_date}"
    else:
        start_date = end_date = PersianDate.today_str()
        title = "سود و زیان"

    txns = get_transactions(start_date=start_date, end_date=end_date)
    expenses = get_expenses(start_date=start_date, end_date=end_date)

    total_income = sum(t.get("final_amount", t["amount"]) for t in txns)
    total_commission = sum(t.get("commission", 0) for t in txns)
    total_tips = sum(t.get("tip", 0) for t in txns)
    total_expenses = sum(e["amount"] for e in expenses)
    net_profit = total_income - total_expenses

    # Expense breakdown
    expense_cats = defaultdict(int)
    for e in expenses:
        expense_cats[e["category"]] += e["amount"]

    # Actual paid salary from expenses (recorded via "ثبت پرداخت حقوق")
    paid_salary = sum(e["amount"] for e in expenses if e["category"] == "حقوق و دستمزد")

    # Income by service category
    income_cats = defaultdict(int)
    for t in txns:
        cat = t.get("category", "") or "نامشخص"
        income_cats[cat] += t.get("final_amount", t["amount"])

    # Payroll for selected period
    payroll = employee_payroll(get_employees(), txns)
    total_salary = sum(p["salary"] for p in payroll.values())
    total_payroll = sum(p["net"] for p in payroll.values())

    return render_template("profit_loss.html",
        title=title, period=period,
        start_date=start_date, end_date=end_date,
        total_income=total_income, total_commission=total_commission,
        total_tips=total_tips, total_expenses=total_expenses,
        net_profit=net_profit,
        expense_cats=dict(expense_cats), income_cats=dict(income_cats),
        txn_count=len(txns), expense_count=len(expenses),
        payroll=payroll, total_salary=total_salary, total_payroll=total_payroll,
        paid_salary=paid_salary)

# ─── Routes: Customer History ───
@app.route("/customer/<name>")
@login_required
def customer_history(name):
    all_txns = get_transactions()
    cust_txns = [t for t in all_txns if t["customer"] == name]
    custs = get_customers()
    cust = next((c for c in custs if c["name"] == name), None)
    total_spent = sum(t.get("final_amount", t["amount"]) for t in cust_txns)
    total_commission = sum(t.get("commission", 0) for t in cust_txns)
    services_used = defaultdict(lambda: {"count":0,"total":0})
    for t in cust_txns:
        services_used[t["service"]]["count"] += 1
        services_used[t["service"]]["total"] += t.get("final_amount", t["amount"])
    return render_template("customer_history.html",
        customer=cust, transactions=list(reversed(cust_txns)),
        total_spent=total_spent, total_commission=total_commission,
        services_used=dict(services_used))

# ─── Routes: Customers ───
@app.route("/customers", methods=["GET","POST"])
@login_required
def customers_page():
    if request.method == "POST":
        action = request.form.get("action")
        custs = get_customers()
        if action == "add":
            custs.append({"name":request.form["name"],"phone":request.form.get("phone",""),"specialty":request.form.get("specialty",""),"join_date":PersianDate.today_str(),"birth_date":request.form.get("birth_date",""),"visit_count":int(request.form.get("visits","0") or 0),"note":request.form.get("note",""),"points":0})
        elif action == "edit":
            idx = int(request.form["idx"])
            custs[idx] = {"name":request.form["name"],"phone":request.form.get("phone",""),"specialty":request.form.get("specialty",""),"join_date":custs[idx]["join_date"],"birth_date":request.form.get("birth_date",custs[idx].get("birth_date","")),"visit_count":int(request.form.get("visits","0") or 0),"note":request.form.get("note",""),"points":custs[idx].get("points",0)}
        elif action == "delete":
            custs.pop(int(request.form["idx"]))
        save_customers(custs)
        return redirect(url_for("customers_page"))
    query = request.args.get("q","").strip().lower()
    custs = get_customers()
    if query:
        custs = [c for c in custs if query in c["name"].lower() or query in c["phone"] or query in c["specialty"].lower()]
    # Birthday customers (today, Jalali)
    today_j = PersianDate.today_str()  # YYYY/MM/DD
    tj = today_j.split("/")
    birthday_today = [c["name"] for c in custs if c.get("birth_date","") and c["birth_date"].split("/")[-2:] == tj[1:]]
    return render_template("customers.html", customers=custs, query=query, birthday_today=birthday_today)

@app.route("/customers/export")
def export_customers():
    custs = get_customers()
    wb = Workbook(); ws = wb.active; ws.title = "دفترچه مشتریان"
    headers = ["نام","تلفن","تخصص مورد علاقه","تاریخ عضویت","تاریخ تولد","تعداد مراجعه","یادداشت","امتیاز"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h); cell.font = HDR_FONT; cell.fill = PINK
    for r, cu in enumerate(custs, 2):
        ws.cell(row=r,column=1,value=cu["name"]); ws.cell(row=r,column=2,value=cu["phone"])
        ws.cell(row=r,column=3,value=cu["specialty"]); ws.cell(row=r,column=4,value=cu["join_date"])
        ws.cell(row=r,column=5,value=cu.get("birth_date","")); ws.cell(row=r,column=6,value=cu["visit_count"])
        ws.cell(row=r,column=7,value=cu["note"]); ws.cell(row=r,column=8,value=cu.get("points",0))
    for c in range(1,8): ws.column_dimensions[get_column_letter(c)].width = 18
    fp = os.path.join(DATA_DIR, f"customers_{PersianDate.today_str().replace('/','-')}.xlsx")
    wb.save(fp)
    return send_file(fp, as_attachment=True)

# ─── Routes: Employees ───
@app.route("/employees", methods=["GET","POST"])
@login_required
def employees_page():
    if request.method == "POST":
        action = request.form.get("action")
        emps = get_employees()
        base = {"name":request.form["name"],"specialty":request.form.get("specialty",""),"phone":request.form.get("phone",""),
                "share_percent":float(request.form.get("share","0") or 0),
                "salary":int(request.form.get("salary","0") or 0),
                "pay_type":request.form.get("pay_type","ماهانه"),
                "start_date":request.form.get("start_date",""),
                "status":request.form.get("status","فعال"),
                "deductions":int(request.form.get("deductions","0") or 0),
                "deduction_note":request.form.get("deduction_note","")}
        if action == "add":
            emps.append(base)
        elif action == "edit":
            emps[int(request.form["idx"])] = base
        elif action == "delete":
            emps.pop(int(request.form["idx"]))
        save_employees(emps)
        return redirect(url_for("employees_page"))
    return render_template("employees.html", employees=get_employees())

@app.route("/payroll")
@login_required
def payroll_page():
    # Month selector (Jalali)
    now = datetime.now()
    jy, jm, jd = PersianDate.gregorian_to_jalali(now.year, now.month, now.day)
    selected = request.args.get("month", f"{jy}/{jm:02d}")
    try:
        sy, sm = map(int, selected.split("/"))
    except:
        sy, sm = jy, jm
    # month range
    start = f"{sy}/{sm:02d}/01"
    end = f"{sy+1}/01/01" if sm == 12 else f"{sy}/{sm+1:02d}/01"
    txns = get_transactions(start_date=start, end_date=end)
    emps = get_employees()
    pr = employee_payroll(emps, txns, month=selected)
    deductions_list = get_deductions(month=selected)
    total_gross = sum(v["gross"] for v in pr.values())
    total_ded = sum(v["deductions"] for v in pr.values())
    total_net = sum(v["net"] for v in pr.values())
    total_commission = sum(v["commission"] for v in pr.values())
    # month list for selector
    months = []
    for yy in range(jy-1, jy+1):
        for mm in range(1, 13):
            months.append(f"{yy}/{mm:02d}")
    return render_template("payroll.html",
        selected=selected, months=months,
        payroll=pr, total_gross=total_gross, total_ded=total_ded, total_net=total_net, total_commission=total_commission,
        deductions_list=deductions_list,
        employee_transactions={e["name"]: [t for t in txns if t["employee"]==e["name"]] for e in emps})

@app.route("/payroll/deduction", methods=["POST"])
@login_required
def payroll_add_deduction():
    emp = request.form.get("employee", "")
    dtype = request.form.get("dtype", "سایر")
    amount = int(request.form.get("amount", 0) or 0)
    month = request.form.get("month", "")
    if not month:
        jy, jm, jd = PersianDate.gregorian_to_jalali(datetime.now().year, datetime.now().month, datetime.now().day)
        month = f"{jy}/{jm:02d}"
    note = request.form.get("note", "")
    today = PersianDate.today_str()
    if emp and amount > 0:
        add_deduction(emp, dtype, amount, today, month, note)
        flash(f"✅ کسورات «{dtype}» برای {emp} ثبت شد ({amount:,} تومان)", "success")
    else:
        flash("لطفاً کارمند، مبلغ و ماه را وارد کنید", "error")
    return redirect(url_for("payroll_page", month=month))

@app.route("/payroll/deduction/delete", methods=["POST"])
@login_required
def payroll_delete_deduction():
    idx = int(request.form.get("idx", -1))
    month = request.form.get("month", "")
    delete_deduction(idx)
    flash("🗑️ کسورات حذف شد", "info")
    return redirect(url_for("payroll_page", month=month))

@app.route("/payroll/pay", methods=["POST"])
@login_required
def payroll_pay():
    today = PersianDate.today_str()
    emps = get_employees()
    active = [e for e in emps if e.get("status", "فعال") == "فعال" and e.get("salary", 0) > 0]
    if not active:
        flash("کارمند فعالی با حقوق ثابت وجود ندارد", "info")
        return redirect(url_for("employees_page"))
    # Compute net per employee for the current month using real payroll
    jy, jm, jd = PersianDate.gregorian_to_jalali(datetime.now().year, datetime.now().month, datetime.now().day)
    month_str = f"{jy}/{jm:02d}"
    start = f"{jy}/{jm:02d}/01"
    end = f"{jy+1}/01/01" if jm == 12 else f"{jy}/{jm+1:02d}/01"
    month_txns = get_transactions(start_date=start, end_date=end)
    pr = employee_payroll(emps, month_txns, month=month_str)
    total = 0
    for e in active:
        net = pr[e["name"]]["net"]
        total += net
        note = f"پرداخت حقوق {e['name']} — {e.get('pay_type','ماهانه')} | حقوق: {e['salary']:,} + پورسانت: {pr[e['name']]['commission']:,} - کسورات: {pr[e['name']]['deductions']:,}"
        add_expense(today, "حقوق و دستمزد", net, note, "نقدی", "ثبت خودکار از پرونده حقوق")
    flash(f"✅ حقوق {len(active)} کارمند ثبت شد (مجموع خالص {total:,} تومان)", "success")
    return redirect(url_for("payroll_page"))

# ─── Routes: Services ───
@app.route("/services", methods=["GET","POST"])
def services_page():
    if request.method == "POST":
        action = request.form.get("action")
        svcs = get_services()
        if action == "add":
            svcs.append({"name":request.form["name"],"category":request.form.get("category",""),"default_price":int(request.form.get("price","0") or 0)})
        elif action == "edit":
            svcs[int(request.form["idx"])] = {"name":request.form["name"],"category":request.form.get("category",""),"default_price":int(request.form.get("price","0") or 0)}
        elif action == "delete":
            svcs.pop(int(request.form["idx"]))
        save_services(svcs)
        return redirect(url_for("services_page"))
    return render_template("services.html", services=get_services())

# ─── Routes: Backup ───
@app.route("/backup", methods=["GET","POST"])
@login_required
def backup_page():
    if request.method == "POST":
        action = request.form.get("action")
        today = PersianDate.today_str()
        all_txns = get_transactions()
        if action == "daily":
            txns = get_transactions(start_date=today, end_date=today)
            if txns:
                fp = create_backup_excel(txns, f"backup_daily_{today.replace('/','-')}.xlsx")
                flash(f"✅ بکاپ روزانه ذخیره شد! ({len(txns)} تراکنش)","success")
            else: flash("تراکنشی برای امروز نیست","info")
        elif action == "weekly":
            now = datetime.now(); week_ago = now - timedelta(days=7)
            s_jy,s_jm,s_jd = PersianDate.gregorian_to_jalali(week_ago.year,week_ago.month,week_ago.day)
            e_jy,e_jm,e_jd = PersianDate.gregorian_to_jalali(now.year,now.month,now.day)
            sd = f"{s_jy}/{s_jm:02d}/{s_jd:02d}"; ed = f"{e_jy}/{e_jm:02d}/{e_jd:02d}"
            txns = [t for t in all_txns if sd <= t["date"] <= ed]
            if txns:
                fp = create_backup_excel(txns, f"backup_weekly_{ed.replace('/','-')}.xlsx")
                flash(f"✅ بکاپ هفتگی! ({len(txns)} تراکنش)","success")
            else: flash("تراکنشی در ۷ روز اخیر نیست","info")
        elif action == "monthly":
            now = datetime.now(); jy,jm,jd = PersianDate.gregorian_to_jalali(now.year,now.month,now.day)
            sd = f"{jy}/{jm:02d}/01"; ed = f"{jy}/{jm+1:02d}/01" if jm<12 else f"{jy+1}/01/01"
            txns = [t for t in all_txns if sd <= t["date"] < ed]
            if txns:
                fp = create_backup_excel(txns, f"backup_monthly_{jy}_{jm:02d}.xlsx")
                flash(f"✅ بکاپ ماهانه! ({len(txns)} تراکنش)","success")
            else: flash("تراکنشی برای این ماه نیست","info")
        elif action == "zip":
            import zipfile as _zip
            stamp = PersianDate.today_str().replace("/", "-")
            zpath = os.path.join(BASE_DIR, f"backup_full_{stamp}.zip")
            with _zip.ZipFile(zpath, "w", _zip.ZIP_DEFLATED) as z:
                for root, _, files in os.walk(DATA_DIR):
                    for fl in files:
                        fp = os.path.join(root, fl)
                        z.write(fp, os.path.relpath(fp, BASE_DIR))
            from flask import send_file
            return send_file(zpath, as_attachment=True, download_name=f"heli_backup_{stamp}.zip")
        elif action == "full":
            if all_txns:
                fp = create_backup_excel(all_txns, f"backup_full_{today.replace('/','-')}.xlsx")
                flash(f"✅ بکاپ کامل! ({len(all_txns)} تراکنش)","success")
            else: flash("هیچ تراکنشی وجود ندارد","info")
        elif action == "custom":
            fd = request.form.get("from_date",""); td = request.form.get("to_date","")
            txns = [t for t in all_txns if fd <= t["date"] <= td]
            if txns:
                fn = f"backup_custom_{fd.replace('/','-')}_to_{td.replace('/','-')}.xlsx"
                fp = create_backup_excel(txns, fn)
                flash(f"✅ بکاپ ذخیره شد! ({len(txns)} تراکنش)","success")
            else: flash("تراکنشی در این بازه نیست","info")
        return redirect(url_for("backup_page"))
    backups = []
    if os.path.exists(DATA_DIR):
        for f in sorted(os.listdir(DATA_DIR), reverse=True):
            if f.startswith("backup_") and f.endswith(".xlsx"):
                fp = os.path.join(DATA_DIR, f)
                dt = datetime.fromtimestamp(os.path.getmtime(fp))
                jy,jm,jd = PersianDate.gregorian_to_jalali(dt.year,dt.month,dt.day)
                backups.append({"filename":f,"date":f"{jy}/{jm:02d}/{jd:02d} {dt.strftime('%H:%M')}","size":_file_size(fp)})
    return render_template("backup.html", backups=backups[:20], today=PersianDate.today_str())

@app.route("/download/<filename>")
def download(filename):
    fp = os.path.join(DATA_DIR, filename)
    if os.path.exists(fp): return send_file(fp, as_attachment=True)
    flash("فایل یافت نشد","error"); return redirect(url_for("backup_page"))

@app.route("/invoice/<invoice_no>")
@login_required
def view_invoice(invoice_no):
    fp = os.path.join(DATA_DIR, "invoices", f"{invoice_no}.pdf")
    if os.path.exists(fp): return send_file(fp, as_attachment=False, mimetype="application/pdf")
    flash("فاکتور یافت نشد","error"); return redirect(url_for("index"))

@app.route("/invoice/pdf/<invoice_no>")
@login_required
def download_invoice(invoice_no):
    fp = os.path.join(DATA_DIR, "invoices", f"{invoice_no}.pdf")
    if os.path.exists(fp): return send_file(fp, as_attachment=True)
    flash("فاکتور یافت نشد","error"); return redirect(url_for("index"))

# ─── API ───
@app.route("/api/services")
def api_services():
    return jsonify(get_services())

@app.route("/api/employees")
def api_employees():
    return jsonify(get_employees())

@app.route("/api/customers")
def api_customers():
    return jsonify(get_customers())

@app.route("/api/customers/add", methods=["POST"])
@login_required
def api_customers_add():
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    phone = (data.get("phone") or "").strip()
    specialty = (data.get("specialty") or "").strip()
    birth_date = (data.get("birth_date") or "").strip()
    note = (data.get("note") or "").strip()
    if not name:
        return jsonify({"ok": False, "error": "نام مشتری الزامی است"}), 400
    custs = get_customers()
    if any(c["name"] == name for c in custs):
        return jsonify({"ok": False, "error": "این مشتری قبلاً ثبت شده"}), 409
    custs.append({
        "name": name, "phone": phone, "specialty": specialty,
        "join_date": PersianDate.today_str(), "birth_date": birth_date,
        "visit_count": 0, "note": note, "points": 0
    })
    save_customers(custs)
    return jsonify({"ok": True, "name": name})

@app.route("/api/calculate_commission", methods=["POST"])
def api_calculate_commission():
    data = request.json
    emp_name = data.get("employee", "")
    amount = int(data.get("amount", 0))
    emps = get_employees()
    emp = next((e for e in emps if e["name"] == emp_name), None)
    if emp:
        commission = int(amount * emp["share_percent"] / 100)
        salon_share = amount - commission
        return jsonify({"commission": commission, "salon_share": salon_share, "percent": emp["share_percent"]})
    return jsonify({"commission": 0, "salon_share": amount, "percent": 0})

# Initialize auth DB (create users table + default admin) at import time so it
# works under Gunicorn (which imports web_app:app directly, not run.py).
try:
    auth.init_auth()
except Exception as _e:
    import traceback as _tb
    print("init_auth failed:", _e)
    _tb.print_exc()

# ─── Run ───
if __name__ == "__main__":
    init_all()
    app.run(host="0.0.0.0", port=5000, debug=True)
